import re
from pathlib import Path

import openpyxl


def ReadLine(sFileName, sTestID):
    path = str(Path(__file__).parent.parent.parent) + "/testResources/TestData" + "/" + sFileName
    wb = openpyxl.load_workbook(path, True)
    wSheet = wb.worksheets.__getitem__(0)
    currentRow = {}
    rowNumber = 0
    for row in range(wSheet.max_row):
        scellValue = wSheet.cell(row + 1, 1).value
        if len(re.findall(scellValue, sTestID)) > 0:
            rowNumber = wSheet.cell(row, 1).row
            break

    for col in range(wSheet.max_column):
        currentRow[wSheet.cell(1, col + 1).value] = wSheet.cell(rowNumber + 1, col + 1).value
    return currentRow


def getLastRow(wb):
    return wb.wsheet.max_row
